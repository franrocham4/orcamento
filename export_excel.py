"""
Módulo para exportar dados em Excel
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from typing import List, Dict, Any
import logging

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Exportador de dados para Excel"""

    def __init__(self):
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def export_company_expenses(self, company_name: str, company_code: str, 
                               contract_value: float, spent_value: float,
                               expenses: List[Dict[str, Any]]) -> str:
        """
        Exporta lançamentos de uma empresa para Excel
        
        Args:
            company_name: Nome da empresa
            company_code: Código da empresa
            contract_value: Valor total do contrato
            spent_value: Valor gasto
            expenses: Lista de lançamentos
            
        Returns:
            Caminho do arquivo gerado
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Movimentos"

            # Configurar largura das colunas
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 20

            # Cabeçalho com informações da empresa
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)

            ws['A1'] = "RELATÓRIO DE MOVIMENTOS"
            ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
            ws.merge_cells('A1:F1')

            ws['A2'] = f"Empresa: {company_name}"
            ws['A2'].font = Font(bold=True, size=11)
            ws.merge_cells('A2:F2')

            ws['A3'] = f"Código: {company_code}"
            ws['A3'].font = Font(size=10)
            ws.merge_cells('A3:F3')

            # Informações financeiras
            ws['A4'] = "Data do Relatório:"
            ws['B4'] = datetime.now().strftime('%d/%m/%Y %H:%M')
            ws['A4'].font = Font(bold=True)

            ws['A5'] = "Valor do Contrato:"
            ws['B5'] = contract_value
            ws['B5'].number_format = 'R$ #,##0.00'
            ws['A5'].font = Font(bold=True)

            ws['A6'] = "Valor Gasto:"
            ws['B6'] = spent_value
            ws['B6'].number_format = 'R$ #,##0.00'
            ws['A6'].font = Font(bold=True)

            available = contract_value - spent_value
            ws['A7'] = "Valor Disponível:"
            ws['B7'] = available
            ws['B7'].number_format = 'R$ #,##0.00'
            ws['A7'].font = Font(bold=True)

            percentage = (spent_value / contract_value * 100) if contract_value > 0 else 0
            ws['A8'] = "Percentual Utilizado:"
            ws['B8'] = percentage / 100
            ws['B8'].number_format = '0.00%'
            ws['A8'].font = Font(bold=True)

            # Cabecalho da tabela
            row = 10
            headers = ['Data', 'Descricao', 'Categoria', 'Valor', 'Quem Registrou', 'Observacoes', 'Data de Criacao']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.thin_border

            # Dados dos lancamentos
            row = 11
            for expense in expenses:
                ws.cell(row=row, column=1).value = expense.get('expense_date', '')
                ws.cell(row=row, column=2).value = expense.get('description', '')
                ws.cell(row=row, column=3).value = expense.get('category', '')
                ws.cell(row=row, column=4).value = expense.get('amount', 0)
                ws.cell(row=row, column=5).value = expense.get('created_by', 'N/A')
                ws.cell(row=row, column=6).value = expense.get('notes', '')
                ws.cell(row=row, column=7).value = expense.get('created_at', '')

                # Formatar valores monetários
                ws.cell(row=row, column=4).number_format = 'R$ #,##0.00'

                # Aplicar bordas
                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = self.thin_border
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='left', vertical='center')

                row += 1

            # Rodapé com totalizações
            if expenses:
                total_row = row + 1
                ws.cell(row=total_row, column=2).value = "TOTAL"
                ws.cell(row=total_row, column=2).font = Font(bold=True)
                ws.cell(row=total_row, column=4).value = sum(e.get('amount', 0) for e in expenses)
                ws.cell(row=total_row, column=4).font = Font(bold=True)
                ws.cell(row=total_row, column=4).number_format = 'R$ #,##0.00'

                for col in range(1, 7):
                    ws.cell(row=total_row, column=col).border = self.thin_border
                    ws.cell(row=total_row, column=col).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

            # Salvar arquivo
            filename = f"Movimentos_{company_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = f"downloads/{filename}"
            
            # Criar pasta downloads se não existir
            import os
            os.makedirs('downloads', exist_ok=True)
            
            wb.save(filepath)
            logger.info(f"Arquivo exportado: {filepath}")
            
            return filepath

        except Exception as e:
            logger.error(f"Erro ao exportar para Excel: {e}")
            import traceback
            traceback.print_exc()
            return None
