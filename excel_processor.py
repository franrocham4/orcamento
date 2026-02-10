"""
Processador de Excel sem dependência de pandas
Usa apenas openpyxl que é mais leve
"""

from openpyxl import load_workbook
from typing import List, Dict, Any
from pathlib import Path
import logging

logger = logging.getLogger(__name__)


class CompanyData:
    """Classe para armazenar dados de uma empresa"""
    def __init__(self, code: str, name: str, contract_value: float, spent_value: float):
        self.code = code
        self.name = name
        self.contract_value = contract_value
        self.spent_value = spent_value

    def to_dict(self) -> Dict[str, Any]:
        """Converte para dicionário"""
        percentage = (self.spent_value / self.contract_value * 100) if self.contract_value > 0 else 0
        return {
            'code': self.code,
            'name': self.name,
            'contract_value': self.contract_value,
            'spent_value': self.spent_value,
            'percentage': round(percentage, 2),
            'status': self._get_status(percentage)
        }

    @staticmethod
    def _get_status(percentage: float) -> str:
        """Retorna status baseado no percentual"""
        if percentage > 90:
            return 'critical'
        elif percentage > 70:
            return 'warning'
        else:
            return 'ok'


class ExcelProcessor:
    """Processador de arquivos Excel usando openpyxl"""

    def __init__(self):
        self.companies: Dict[str, CompanyData] = {}
        self.last_data = {
            'companies': [],
            'statistics': {}
        }

    def process_file(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Processa arquivo Excel e retorna lista de empresas

        Args:
            file_path: Caminho do arquivo Excel

        Returns:
            Lista de dicionários com dados das empresas
        """
        try:
            if not Path(file_path).exists():
                logger.error(f"Arquivo não encontrado: {file_path}")
                return []

            # Carregar workbook
            wb = load_workbook(file_path, data_only=True)

            # Verificar se as abas existem
            sheet_names = wb.sheetnames
            
            if 'VALIDAÇÕES' not in sheet_names:
                logger.error(f"Aba 'VALIDAÇÕES' não encontrada. Abas disponíveis: {sheet_names}")
                return []

            if 'LIQUIDAÇÃO 2025' not in sheet_names:
                logger.error(f"Aba 'LIQUIDAÇÃO 2025' não encontrada. Abas disponíveis: {sheet_names}")
                return []

            # Processar abas
            self._process_validacoes(wb['VALIDAÇÕES'])
            self._process_liquidacao(wb['LIQUIDAÇÃO 2025'])

            # Converter para lista de dicionários
            result = [company.to_dict() for company in self.companies.values()]

            # Ordenar por nome
            result.sort(key=lambda x: x['name'])

            logger.info(f"Processadas {len(result)} empresas")
            return result

        except Exception as e:
            logger.error(f"Erro ao processar arquivo: {e}")
            import traceback
            traceback.print_exc()
            return []

    def _process_validacoes(self, ws) -> None:
        """Processa aba VALIDAÇÕES"""
        try:
            logger.info("Processando aba VALIDAÇÕES...")
            
            # Iterar sobre as linhas
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row_idx == 1:  # Pular cabeçalho
                    continue
                
                if not row or len(row) < 7:
                    continue

                # Colunas: 0=Código, 1=Empresa, 6=Valor Contrato
                codigo = str(row[0]).strip() if row[0] else ""
                empresa = str(row[1]).strip() if row[1] else ""
                valor = row[6] if row[6] else 0

                # Pular linhas vazias
                if not codigo or not empresa:
                    continue

                try:
                    valor_float = float(valor) if isinstance(valor, (int, float)) else 0
                    if valor_float > 0:  # Só adicionar se tiver valor
                        self.companies[codigo] = CompanyData(codigo, empresa, valor_float, 0)
                        logger.debug(f"Empresa adicionada: {codigo} - {empresa} - R${valor_float}")
                except (ValueError, TypeError) as e:
                    logger.debug(f"Erro ao processar valor: {e}")
                    continue

            logger.info(f"Total de empresas após VALIDAÇÕES: {len(self.companies)}")

        except Exception as e:
            logger.error(f"Erro ao processar VALIDAÇÕES: {e}")
            import traceback
            traceback.print_exc()

    def _process_liquidacao(self, ws) -> None:
        """Processa aba LIQUIDAÇÃO 2025"""
        try:
            logger.info("Processando aba LIQUIDAÇÃO 2025...")
            
            # Dicionário para acumular gastos por código
            gastos_por_codigo = {}

            # Iterar sobre as linhas
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row_idx == 1:  # Pular cabeçalho
                    continue
                
                if not row or len(row) < 7:
                    continue

                # Colunas: 1=Código, 6=Valor Liquidado
                codigo = str(row[1]).strip() if row[1] else ""
                valor = row[6] if row[6] else 0

                # Pular linhas vazias
                if not codigo:
                    continue

                try:
                    valor_float = float(valor) if isinstance(valor, (int, float)) else 0
                    if valor_float > 0:
                        if codigo not in gastos_por_codigo:
                            gastos_por_codigo[codigo] = 0
                        gastos_por_codigo[codigo] += valor_float
                        logger.debug(f"Gasto registrado: {codigo} - R${valor_float}")
                except (ValueError, TypeError) as e:
                    logger.debug(f"Erro ao processar valor: {e}")
                    continue

            # Atualizar gastos nas empresas
            for codigo, gasto in gastos_por_codigo.items():
                if codigo in self.companies:
                    company = self.companies[codigo]
                    company.spent_value = gasto
                    logger.debug(f"Gasto atualizado para {codigo}: R${gasto}")

            logger.info(f"Total de empresas após LIQUIDAÇÃO: {len(self.companies)}")

        except Exception as e:
            logger.error(f"Erro ao processar LIQUIDAÇÃO 2025: {e}")
            import traceback
            traceback.print_exc()

    def get_statistics(self, companies: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Calcula estatísticas gerais

        Args:
            companies: Lista de empresas

        Returns:
            Dicionário com estatísticas
        """
        if not companies:
            return {
                'total_contracted': 0,
                'total_spent': 0,
                'average_utilization': 0,
                'companies_count': 0
            }

        total_contracted = sum(c['contract_value'] for c in companies)
        total_spent = sum(c['spent_value'] for c in companies)
        average_utilization = (total_spent / total_contracted * 100) if total_contracted > 0 else 0

        return {
            'total_contracted': total_contracted,
            'total_spent': total_spent,
            'average_utilization': round(average_utilization, 2),
            'companies_count': len(companies)
        }

    def get_data(self) -> Dict[str, Any]:
        """Retorna os dados atuais processados"""
        companies = [company.to_dict() for company in self.companies.values()]
        companies.sort(key=lambda x: x['name'])
        
        return {
            'companies': companies,
            'statistics': self.get_statistics(companies)
        }
